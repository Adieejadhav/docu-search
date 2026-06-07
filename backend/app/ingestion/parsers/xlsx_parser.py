from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.constants import BlockType, SupportedFileType, SUPPORTED_XLSX_EXTENSIONS
from app.core.exceptions import ParserError
from app.ingestion.normalizers.block_schema import (
    DocumentBlock,
    ParsedDocument,
    SourceLocation,
)
from app.ingestion.parsers._tabular import (
    dedupe_headers,
    format_key_value_row,
    normalize_row,
    row_is_empty,
)
from app.ingestion.parsers.base import BaseDocumentParser


class XlsxParser(BaseDocumentParser):
    """
    Parser for Excel .xlsx workbooks.

    Each non-empty worksheet becomes a heading block. The first non-empty row
    in that sheet is treated as the header row, and following rows become table
    blocks.
    """

    supported_extensions = SUPPORTED_XLSX_EXTENSIONS

    def parse(self, file_path: str | Path) -> ParsedDocument:
        path = self.validate_input_file(file_path)
        workbook = self._open_workbook(path)

        try:
            blocks = self._parse_workbook(workbook)
            title = workbook.properties.title or self.get_document_title(path)
        finally:
            workbook.close()

        return self.build_document(
            path=path,
            title=title,
            file_type=SupportedFileType.XLSX,
            blocks=blocks,
            metadata={"sheet_count": len(workbook.sheetnames)},
        )

    def _open_workbook(self, path: Path) -> Any:
        try:
            return load_workbook(
                filename=str(path),
                read_only=True,
                data_only=True,
            )
        except InvalidFileException as exc:
            raise ParserError(
                f"Unable to open XLSX file: {path}",
                code="XLSX_OPEN_FAILED",
                details={"file_path": str(path)},
            ) from exc
        except Exception as exc:
            raise ParserError(
                f"Unable to parse XLSX file: {path}",
                code="XLSX_PARSE_FAILED",
                details={"file_path": str(path)},
            ) from exc

    def _parse_workbook(self, workbook: Any) -> list[DocumentBlock]:
        blocks: list[DocumentBlock] = []

        for worksheet in workbook.worksheets:
            sheet_blocks = self._parse_worksheet(worksheet, start_order=len(blocks))
            blocks.extend(sheet_blocks)

        return blocks

    def _parse_worksheet(
        self, worksheet: Any, *, start_order: int
    ) -> list[DocumentBlock]:
        blocks: list[DocumentBlock] = []
        headers: list[str] | None = None
        header_row_number: int | None = None

        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            normalized_row = normalize_row(row)
            if row_is_empty(normalized_row):
                continue

            if headers is None:
                headers = dedupe_headers(normalized_row)
                header_row_number = row_number
                blocks.append(
                    DocumentBlock(
                        block_type=BlockType.HEADING,
                        text=worksheet.title,
                        order=start_order + len(blocks),
                        level=1,
                        source_location=SourceLocation(sheet_name=worksheet.title),
                        metadata={
                            "sheet_name": worksheet.title,
                            "header_row": header_row_number,
                        },
                    )
                )
                continue

            text = format_key_value_row(headers, normalized_row)
            if not text:
                continue

            blocks.append(
                DocumentBlock(
                    block_type=BlockType.TABLE,
                    text=text,
                    order=start_order + len(blocks),
                    parent_path=[worksheet.title],
                    source_location=SourceLocation(
                        sheet_name=worksheet.title,
                        row_start=row_number,
                        row_end=row_number,
                    ),
                    metadata={
                        "format": "xlsx_row",
                        "sheet_name": worksheet.title,
                        "header_row": header_row_number,
                        "row_number": row_number,
                        "columns": headers,
                    },
                )
            )

        if headers is not None and len(blocks) == 1:
            blocks.append(
                DocumentBlock(
                    block_type=BlockType.TABLE,
                    text="Columns: " + " | ".join(headers),
                    order=start_order + len(blocks),
                    parent_path=[worksheet.title],
                    source_location=SourceLocation(
                        sheet_name=worksheet.title,
                        row_start=header_row_number,
                        row_end=header_row_number,
                    ),
                    metadata={
                        "format": "xlsx_header",
                        "sheet_name": worksheet.title,
                        "header_row": header_row_number,
                        "columns": headers,
                    },
                )
            )

        return blocks
